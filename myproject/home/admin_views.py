"""
Custom admin CRUD views. All require admin_required (role Admin or Django staff/superuser).
"""
from django.urls import reverse, reverse_lazy
from django.views.generic import ListView, CreateView, UpdateView, DeleteView, TemplateView
from django.db.models import Q
from django.db import transaction
from django.utils.decorators import method_decorator
from django.contrib import messages
from django.shortcuts import redirect
from django.core.serializers import serialize
from django.http import HttpResponse
from django.db import connection
from django.db.utils import OperationalError, ProgrammingError
import json

from .decorators import admin_required
from .admin_mixins import AdminSaveRedirectMixin, AdminBulkDeleteMixin
from .models import (
    Building,
    UniversityBranch,
    Floor,
    Room,
    Tree,
    Equipment,
    Incident,
    IncidentType,
    Asset,
    Maintenance,
    Role,
    AppUser,
    AboutHeroSection,
    AboutCoreValue,
    AboutAnnouncement,
    AboutFeaturedEvent,
)
from .map_utils import buildings_geojson_dumps
from .admin_forms import (
    UniversityBranchAdminForm,
    TreeAdminForm,
    EquipmentAdminForm,
    RoomAdminForm,
    BuildingAdminForm,
    FloorAdminForm,
    IncidentAdminForm,
    IncidentTypeAdminForm,
    AssetAdminForm,
    MaintenanceAdminForm,
    RoleAdminForm,
    AppUserAdminForm,
    AboutHeroSectionAdminForm,
    AboutCoreValueAdminForm,
    AboutAnnouncementAdminForm,
    AboutFeaturedEventAdminForm,
)
from .admin_excel_import import TABLE_IMPORT_CONFIG, import_rows_from_workbook

try:
    from openpyxl import load_workbook
    from openpyxl import Workbook
except ImportError:  # pragma: no cover - safety for missing dependency
    load_workbook = None
    Workbook = None


def _rooms_admin_overlay_geojson() -> str:
    """GeoJSON của mọi phòng có polygon — dùng trên form admin để lọc theo tầng."""
    features = []
    for r in Room.objects.exclude(geom__isnull=True).select_related("floor", "floor__building"):
        features.append(
            {
                "type": "Feature",
                "geometry": json.loads(r.geom.geojson),
                "properties": {
                    "id": r.id,
                    "name": r.name or "",
                    "floor_id": r.floor_id,
                    "building_name": r.floor.building.name if r.floor and r.floor.building else "",
                },
            }
        )
    return json.dumps({"type": "FeatureCollection", "features": features})


@method_decorator(admin_required, name="dispatch")
class AdminExcelImportView(TemplateView):
    template_name = "home/admin/excel_import.html"

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        table_name = (self.request.GET.get("table") or self.request.POST.get("table") or "equipment").strip().lower()
        if table_name not in TABLE_IMPORT_CONFIG:
            table_name = "equipment"
        ctx["selected_table"] = table_name
        ctx["table_config"] = TABLE_IMPORT_CONFIG
        ctx["table_labels"] = {
            "equipment": "Thiết bị",
            "tree": "Cây xanh",
            "incident": "Sự cố",
            "maintenance": "Bảo trì",
        }
        return ctx

    def post(self, request, *args, **kwargs):
        table_name = (request.POST.get("table") or "").strip().lower()
        if table_name not in TABLE_IMPORT_CONFIG:
            messages.error(request, "Bảng nhập không hợp lệ.")
            return redirect(reverse("admin_excel_import"))

        upload_file = request.FILES.get("excel_file")
        if not upload_file:
            messages.error(request, "Vui lòng chọn file Excel (.xlsx).")
            return redirect(f"{reverse('admin_excel_import')}?table={table_name}")
        if not upload_file.name.lower().endswith(".xlsx"):
            messages.error(request, "Chỉ hỗ trợ file .xlsx.")
            return redirect(f"{reverse('admin_excel_import')}?table={table_name}")
        if load_workbook is None:
            messages.error(request, "Thiếu thư viện openpyxl. Hãy cài đặt openpyxl để dùng tính năng nhập Excel.")
            return redirect(f"{reverse('admin_excel_import')}?table={table_name}")

        workbook = load_workbook(upload_file, data_only=True)
        result = import_rows_from_workbook(table_name, workbook)
        if result.errors:
            messages.error(request, f"Nhập thất bại: {len(result.errors)} lỗi.")
            for err in result.errors[:20]:
                messages.warning(request, err)
            if len(result.errors) > 20:
                messages.warning(request, f"... còn {len(result.errors) - 20} lỗi khác.")
        else:
            messages.success(request, f"Nhập thành công {result.created_count} dòng vào bảng {table_name}.")
        return redirect(f"{reverse('admin_excel_import')}?table={table_name}")


@method_decorator(admin_required, name="dispatch")
class AdminExcelTemplateDownloadView(TemplateView):
    def get(self, request, *args, **kwargs):
        if Workbook is None:
            messages.error(request, "Thiếu thư viện openpyxl, chưa thể tạo file mẫu.")
            return redirect(reverse("admin_excel_import"))

        workbook = Workbook()
        default_sheet = workbook.active
        workbook.remove(default_sheet)

        template_rows = {
            "equipment": [
                "EQ-001", "Máy chiếu Sony", "Projector", "good", "Phòng Lab 501", "2024-01-15", "", "", "", "", ""
            ],
            "tree": [
                "TREE-LVS-001", "Cây bàng", "12.5", "good", "2021-03-20", "", "Trồng mới", "Cơ sở Lê Văn Sỹ", "", ""
            ],
            "incident": [
                "Máy chiếu phòng 501 lỗi", "Màn hình chớp liên tục", "open", "medium", "", "", "EQ-001", "", "", "", "", "", ""
            ],
            "maintenance": [
                "PC-01", "kthuat01", "inspection", "2026-04-13", "50000", "Kiểm tra định kỳ"
            ],
        }

        for table_name, cfg in TABLE_IMPORT_CONFIG.items():
            ws = workbook.create_sheet(title=table_name)
            ws.append(cfg["columns"])
            ws.append(template_rows.get(table_name, [""] * len(cfg["columns"])))

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = 'attachment; filename="mau_import_du_lieu.xlsx"'
        workbook.save(response)
        return response


# ----- University Branch -----
@method_decorator(admin_required, name="dispatch")
class UniversityBranchListView(AdminBulkDeleteMixin, ListView):
    model = UniversityBranch
    template_name = "home/admin/branch_list.html"
    context_object_name = "branches"
    paginate_by = 10
    bulk_delete_message = "Đã xóa chi nhánh đã chọn."

    def get_queryset(self):
        qs = UniversityBranch.objects.all().order_by("name")
        q = (self.request.GET.get("q") or "").strip()
        if q:
            qs = qs.filter(Q(name__icontains=q) | Q(description__icontains=q))
        return qs

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["q"] = self.request.GET.get("q", "")
        return ctx


@method_decorator(admin_required, name="dispatch")
class UniversityBranchCreateView(AdminSaveRedirectMixin, CreateView):
    model = UniversityBranch
    form_class = UniversityBranchAdminForm
    template_name = "home/admin/branch_form.html"
    list_url_name = "admin_branch_list"
    add_url_name = "admin_branch_add"
    edit_url_name = "admin_branch_edit"

    def form_valid(self, form):
        messages.success(self.request, "Đã thêm chi nhánh mới.")
        return super().form_valid(form)


@method_decorator(admin_required, name="dispatch")
class UniversityBranchUpdateView(AdminSaveRedirectMixin, UpdateView):
    model = UniversityBranch
    form_class = UniversityBranchAdminForm
    template_name = "home/admin/branch_form.html"
    list_url_name = "admin_branch_list"
    add_url_name = "admin_branch_add"
    edit_url_name = "admin_branch_edit"

    def form_valid(self, form):
        messages.success(self.request, "Đã cập nhật chi nhánh.")
        return super().form_valid(form)


@method_decorator(admin_required, name="dispatch")
class UniversityBranchDeleteView(DeleteView):
    model = UniversityBranch
    template_name = "home/admin/confirm_delete.html"
    success_url = reverse_lazy("admin_branch_list")
    context_object_name = "obj"

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["model_label"] = "Chi nhánh"
        ctx["list_url"] = "admin_branch_list"
        return ctx

    def delete(self, request, *args, **kwargs):
        messages.success(request, "Đã xóa chi nhánh.")
        return super().delete(request, *args, **kwargs)


def _admin_list_view(model, template_name, context_object_name, order_by, filters=None):
    """Helper to avoid repeating get_queryset logic."""
    class _ListView(ListView):
        model = model
        template_name = template_name
        context_object_name = context_object_name
        paginate_by = 10

        def get_queryset(self):
            qs = super().get_queryset().order_by(order_by)
            if filters:
                for key, q_param in filters.items():
                    val = self.request.GET.get(key)
                    if val:
                        qs = qs.filter(**{q_param: val})
            q = self.request.GET.get("q")
            if q and hasattr(model, "code"):
                qs = qs.filter(Q(code__icontains=q) | Q(name__icontains=q) if hasattr(model, "name") else qs.filter(code__icontains=q))
            elif q:
                if hasattr(model, "name"):
                    qs = qs.filter(name__icontains=q)
                if hasattr(model, "title"):
                    qs = qs.filter(Q(title__icontains=q) | Q(description__icontains=q))
            return qs

        def get_context_data(self, **kwargs):
            ctx = super().get_context_data(**kwargs)
            ctx["q"] = self.request.GET.get("q", "")
            for key in (filters or {}).keys():
                ctx[key] = self.request.GET.get(key, "")
            return ctx

    return _ListView


# ----- Tree -----
@method_decorator(admin_required, name="dispatch")
class TreeListView(AdminBulkDeleteMixin, ListView):
    model = Tree
    template_name = "home/admin/tree_list.html"
    context_object_name = "trees"
    paginate_by = 10
    bulk_delete_message = "Đã xóa cây đã chọn."

    def get_queryset(self):
        qs = Tree.objects.all().order_by("code")
        q = self.request.GET.get("q")
        status = self.request.GET.get("status")
        if q:
            qs = qs.filter(Q(code__icontains=q) | Q(species__icontains=q))
        if status:
            qs = qs.filter(health_status=status)
        return qs

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["q"] = self.request.GET.get("q", "")
        ctx["status_filter"] = self.request.GET.get("status", "")
        return ctx

    def post(self, request, *args, **kwargs):
        action = request.POST.get(self.action_field_name) or ""
        if action != "delete_selected":
            return super().post(request, *args, **kwargs)

        selected = request.POST.getlist(self.selected_field_name)
        if not selected:
            messages.warning(request, "Vui lòng chọn ít nhất 1 dòng để thực hiện hành động.")
            return redirect(request.get_full_path())

        tree_qs = Tree.objects.filter(pk__in=selected)
        deleted_count = tree_qs.count()
        with transaction.atomic():
            Asset.objects.filter(tree_id__in=tree_qs.values_list("id", flat=True)).delete()
            tree_qs.delete()
        messages.success(request, f"{self.bulk_delete_message} (Đã xóa: {deleted_count})")
        return redirect(request.path)


@method_decorator(admin_required, name="dispatch")
class TreeCreateView(AdminSaveRedirectMixin, CreateView):
    model = Tree
    form_class = TreeAdminForm
    template_name = "home/admin/tree_form.html"
    list_url_name = "admin_tree_list"
    add_url_name = "admin_tree_add"
    edit_url_name = "admin_tree_edit"

    def form_valid(self, form):
        messages.success(self.request, "Đã thêm cây xanh mới.")
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["buildings_json"] = serialize(
            "geojson",
            Building.objects.exclude(geom__isnull=True),
            geometry_field="geom",
            fields=("name", "description"),
        )
        ctx["rooms_json"] = serialize(
            "geojson",
            Room.objects.exclude(geom__isnull=True),
            geometry_field="geom",
            fields=("name", "room_type"),
        )
        ctx["trees_json"] = serialize(
            "geojson",
            Tree.objects.exclude(geom__isnull=True),
            geometry_field="geom",
            fields=("code", "species", "health_status"),
        )
        ctx["equipment_json"] = serialize(
            "geojson",
            Equipment.objects.exclude(geom__isnull=True),
            geometry_field="geom",
            fields=("code", "name", "status"),
        )
        ctx["branch_centers_json"] = json.dumps({
            b.id: [b.geom.centroid.y, b.geom.centroid.x]
            for b in UniversityBranch.objects.exclude(geom__isnull=True)
        })
        ctx["buildings_meta_json"] = json.dumps({
            b.id: {"branch_id": b.university_branch_id}
            for b in Building.objects.all()
        })
        ctx["trees_meta_json"] = json.dumps({
            t.id: {"branch_id": t.university_branch_id}
            for t in Tree.objects.all()
        })
        return ctx


@method_decorator(admin_required, name="dispatch")
class TreeUpdateView(AdminSaveRedirectMixin, UpdateView):
    model = Tree
    form_class = TreeAdminForm
    template_name = "home/admin/tree_form.html"
    list_url_name = "admin_tree_list"
    add_url_name = "admin_tree_add"
    edit_url_name = "admin_tree_edit"

    def form_valid(self, form):
        messages.success(self.request, "Đã cập nhật thông tin cây.")
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["buildings_json"] = serialize(
            "geojson",
            Building.objects.exclude(geom__isnull=True),
            geometry_field="geom",
            fields=("name", "description"),
        )
        ctx["rooms_json"] = serialize(
            "geojson",
            Room.objects.exclude(geom__isnull=True),
            geometry_field="geom",
            fields=("name", "room_type"),
        )
        ctx["trees_json"] = serialize(
            "geojson",
            Tree.objects.exclude(geom__isnull=True),
            geometry_field="geom",
            fields=("code", "species", "health_status"),
        )
        ctx["equipment_json"] = serialize(
            "geojson",
            Equipment.objects.exclude(geom__isnull=True),
            geometry_field="geom",
            fields=("code", "name", "status"),
        )
        ctx["branch_centers_json"] = json.dumps({
            b.id: [b.geom.centroid.y, b.geom.centroid.x]
            for b in UniversityBranch.objects.exclude(geom__isnull=True)
        })
        ctx["buildings_meta_json"] = json.dumps({
            b.id: {"branch_id": b.university_branch_id}
            for b in Building.objects.all()
        })
        ctx["trees_meta_json"] = json.dumps({
            t.id: {"branch_id": t.university_branch_id}
            for t in Tree.objects.all()
        })
        return ctx


@method_decorator(admin_required, name="dispatch")
class TreeDeleteView(DeleteView):
    model = Tree
    template_name = "home/admin/confirm_delete.html"
    success_url = reverse_lazy("admin_tree_list")
    context_object_name = "obj"

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["model_label"] = "Cây xanh"
        ctx["list_url"] = "admin_tree_list"
        return ctx

    def form_valid(self, form):
        self.object = self.get_object()
        with transaction.atomic():
            Asset.objects.filter(tree_id=self.object.pk).delete()
            self.object.delete()
        messages.success(self.request, "Đã xóa cây.")
        return redirect(self.success_url)


# ----- Equipment -----
@method_decorator(admin_required, name="dispatch")
class EquipmentListView(AdminBulkDeleteMixin, ListView):
    model = Equipment
    template_name = "home/admin/equipment_list.html"
    context_object_name = "equipment_list"
    paginate_by = 10
    bulk_delete_message = "Đã xóa thiết bị đã chọn."

    def get_queryset(self):
        qs = Equipment.objects.select_related("room", "room__floor", "room__floor__building").order_by("code")
        q = (self.request.GET.get("q") or "").strip()
        status = self.request.GET.get("status")
        if q:
            qs = qs.filter(
                Q(code__icontains=q)
                | Q(name__icontains=q)
                | Q(equipment_type__icontains=q)
                | Q(room__name__icontains=q)
                | Q(room__floor__name__icontains=q)
                | Q(room__floor__building__name__icontains=q)
                | Q(room__floor__building__university_branch__name__icontains=q)
            )
        if status:
            qs = qs.filter(status=status)
        return qs

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["q"] = self.request.GET.get("q", "")
        ctx["status_filter"] = self.request.GET.get("status", "")
        return ctx

    def post(self, request, *args, **kwargs):
        action = request.POST.get(self.action_field_name) or ""
        if action != "delete_selected":
            return super().post(request, *args, **kwargs)

        selected = request.POST.getlist(self.selected_field_name)
        if not selected:
            messages.warning(request, "Vui lòng chọn ít nhất 1 dòng để thực hiện hành động.")
            return redirect(request.get_full_path())

        eq_qs = Equipment.objects.filter(pk__in=selected)
        deleted_count = eq_qs.count()
        with transaction.atomic():
            Asset.objects.filter(equipment__in=eq_qs).delete()
            eq_qs.delete()
        messages.success(request, f"{self.bulk_delete_message} (Đã xóa: {deleted_count})")
        return redirect(request.path)


@method_decorator(admin_required, name="dispatch")
class EquipmentCreateView(AdminSaveRedirectMixin, CreateView):
    model = Equipment
    form_class = EquipmentAdminForm
    template_name = "home/admin/equipment_form.html"
    list_url_name = "admin_equipment_list"
    add_url_name = "admin_equipment_add"
    edit_url_name = "admin_equipment_edit"

    def form_valid(self, form):
        messages.success(self.request, "Đã thêm thiết bị mới.")
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["buildings_json"] = buildings_geojson_dumps()
        ctx["rooms_json"] = serialize(
            "geojson",
            Room.objects.exclude(geom__isnull=True),
            geometry_field="geom",
            fields=("id", "name", "room_type"),
        )
        ctx["trees_json"] = serialize(
            "geojson",
            Tree.objects.exclude(geom__isnull=True),
            geometry_field="geom",
            fields=("code", "species", "health_status"),
        )
        ctx["equipment_json"] = serialize(
            "geojson",
            Equipment.objects.exclude(geom__isnull=True),
            geometry_field="geom",
            fields=("code", "name", "status"),
        )
        rooms_meta = Room.objects.select_related("floor", "floor__building", "floor__building__university_branch").all()
        buildings_meta = Building.objects.select_related("university_branch").all()
        floors_meta = Floor.objects.select_related("building", "building__university_branch").all()
        ctx["buildings_meta_json"] = json.dumps({
            b.id: {"name": b.name, "branch_id": b.university_branch_id} for b in buildings_meta
        })
        ctx["floors_meta_json"] = json.dumps({
            f.id: {"name": f.name, "building_id": f.building_id, "branch_id": (f.building.university_branch_id if f.building else None)}
            for f in floors_meta
        })
        ctx["rooms_meta_json"] = json.dumps({
            r.id: {
                "name": r.name,
                "floor_id": r.floor_id,
                "building_id": (r.floor.building_id if r.floor and r.floor.building_id else None),
                "branch_id": (r.floor.building.university_branch_id if r.floor and r.floor.building else None),
            }
            for r in rooms_meta
        })
        ctx["branch_centers_json"] = json.dumps({
            b.id: [b.geom.centroid.y, b.geom.centroid.x]
            for b in UniversityBranch.objects.exclude(geom__isnull=True)
        })
        return ctx


@method_decorator(admin_required, name="dispatch")
class EquipmentUpdateView(AdminSaveRedirectMixin, UpdateView):
    model = Equipment
    form_class = EquipmentAdminForm
    template_name = "home/admin/equipment_form.html"
    list_url_name = "admin_equipment_list"
    add_url_name = "admin_equipment_add"
    edit_url_name = "admin_equipment_edit"

    def form_valid(self, form):
        messages.success(self.request, "Đã cập nhật thiết bị.")
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["buildings_json"] = buildings_geojson_dumps()
        ctx["rooms_json"] = serialize(
            "geojson",
            Room.objects.exclude(geom__isnull=True),
            geometry_field="geom",
            fields=("id", "name", "room_type"),
        )
        ctx["trees_json"] = serialize(
            "geojson",
            Tree.objects.exclude(geom__isnull=True),
            geometry_field="geom",
            fields=("code", "species", "health_status"),
        )
        ctx["equipment_json"] = serialize(
            "geojson",
            Equipment.objects.exclude(geom__isnull=True),
            geometry_field="geom",
            fields=("code", "name", "status"),
        )
        rooms_meta = Room.objects.select_related("floor", "floor__building", "floor__building__university_branch").all()
        buildings_meta = Building.objects.select_related("university_branch").all()
        floors_meta = Floor.objects.select_related("building", "building__university_branch").all()
        ctx["buildings_meta_json"] = json.dumps({
            b.id: {"name": b.name, "branch_id": b.university_branch_id} for b in buildings_meta
        })
        ctx["floors_meta_json"] = json.dumps({
            f.id: {"name": f.name, "building_id": f.building_id, "branch_id": (f.building.university_branch_id if f.building else None)}
            for f in floors_meta
        })
        ctx["rooms_meta_json"] = json.dumps({
            r.id: {
                "name": r.name,
                "floor_id": r.floor_id,
                "building_id": (r.floor.building_id if r.floor and r.floor.building_id else None),
                "branch_id": (r.floor.building.university_branch_id if r.floor and r.floor.building else None),
            }
            for r in rooms_meta
        })
        ctx["branch_centers_json"] = json.dumps({
            b.id: [b.geom.centroid.y, b.geom.centroid.x]
            for b in UniversityBranch.objects.exclude(geom__isnull=True)
        })
        return ctx


@method_decorator(admin_required, name="dispatch")
class EquipmentDeleteView(DeleteView):
    model = Equipment
    template_name = "home/admin/confirm_delete.html"
    success_url = reverse_lazy("admin_equipment_list")
    context_object_name = "obj"

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["model_label"] = "Thiết bị"
        ctx["list_url"] = "admin_equipment_list"
        return ctx

    def form_valid(self, form):
        self.object = self.get_object()
        with transaction.atomic():
            # Keep explicit cleanup for safety during transition data/migrations.
            Asset.objects.filter(equipment=self.object).delete()
            self.object.delete()
        messages.success(self.request, "Đã xóa thiết bị.")
        return redirect(self.success_url)


# ----- Room -----
@method_decorator(admin_required, name="dispatch")
class RoomListView(AdminBulkDeleteMixin, ListView):
    model = Room
    template_name = "home/admin/room_list.html"
    context_object_name = "rooms"
    paginate_by = 10
    bulk_delete_message = "Đã xóa phòng đã chọn."

    def get_queryset(self):
        qs = Room.objects.select_related("floor", "floor__building").order_by("name")
        q = (self.request.GET.get("q") or "").strip()
        if q:
            filters = Q(name__icontains=q) | Q(floor__name__icontains=q) | Q(floor__building__name__icontains=q)

            # Tìm theo loại phòng bằng tiếng Việt (label choices)
            q_lower = q.lower()
            matching_types = [
                code
                for code, label in Room.ROOM_TYPES
                if q_lower in label.lower()
            ]
            if matching_types:
                filters |= Q(room_type__in=matching_types)

            # Nếu q là số, cho phép tìm theo sức chứa
            if q.isdigit():
                try:
                    capacity_val = int(q)
                    filters |= Q(capacity=capacity_val)
                except ValueError:
                    pass

            qs = qs.filter(filters)
        return qs

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["q"] = self.request.GET.get("q", "")
        return ctx


# ----- Floor -----
@method_decorator(admin_required, name="dispatch")
class FloorListView(AdminBulkDeleteMixin, ListView):
    model = Floor
    template_name = "home/admin/floor_list.html"
    context_object_name = "floors"
    paginate_by = 10
    bulk_delete_message = "Đã xóa tầng đã chọn."

    def get_queryset(self):
        qs = Floor.objects.select_related("building").order_by("building__name", "level")
        q = (self.request.GET.get("q") or "").strip()
        if q:
            qs = qs.filter(Q(name__icontains=q) | Q(building__name__icontains=q) | Q(rooms__name__icontains=q)).distinct()
        return qs

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["q"] = self.request.GET.get("q", "")
        return ctx


@method_decorator(admin_required, name="dispatch")
class FloorCreateView(AdminSaveRedirectMixin, CreateView):
    model = Floor
    form_class = FloorAdminForm
    template_name = "home/admin/floor_form.html"
    list_url_name = "admin_floor_list"
    add_url_name = "admin_floor_add"
    edit_url_name = "admin_floor_edit"

    def form_valid(self, form):
        messages.success(self.request, "Đã thêm tầng mới.")
        return super().form_valid(form)


@method_decorator(admin_required, name="dispatch")
class FloorUpdateView(AdminSaveRedirectMixin, UpdateView):
    model = Floor
    form_class = FloorAdminForm
    template_name = "home/admin/floor_form.html"
    list_url_name = "admin_floor_list"
    add_url_name = "admin_floor_add"
    edit_url_name = "admin_floor_edit"

    def form_valid(self, form):
        messages.success(self.request, "Đã cập nhật thông tin tầng.")
        return super().form_valid(form)


@method_decorator(admin_required, name="dispatch")
class FloorDeleteView(DeleteView):
    model = Floor
    template_name = "home/admin/confirm_delete.html"
    success_url = reverse_lazy("admin_floor_list")
    context_object_name = "obj"

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["model_label"] = "Tầng"
        ctx["list_url"] = "admin_floor_list"
        return ctx

    def delete(self, request, *args, **kwargs):
        messages.success(request, "Đã xóa tầng.")
        return super().delete(request, *args, **kwargs)


@method_decorator(admin_required, name="dispatch")
class RoomCreateView(AdminSaveRedirectMixin, CreateView):
    model = Room
    form_class = RoomAdminForm
    template_name = "home/admin/room_form.html"
    list_url_name = "admin_room_list"
    add_url_name = "admin_room_add"
    edit_url_name = "admin_room_edit"

    def form_valid(self, form):
        messages.success(self.request, "Đã thêm phòng mới.")
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["buildings_json"] = buildings_geojson_dumps()
        floors_meta = Floor.objects.select_related("building", "building__university_branch").all()
        ctx["floors_meta_json"] = json.dumps({
            f.id: {
                "building_id": f.building_id,
                "building_name": (f.building.name if f.building else ""),
                "branch_id": (f.building.university_branch_id if f.building else None),
            }
            for f in floors_meta
        })
        buildings_meta = Building.objects.select_related("university_branch").all()
        ctx["buildings_meta_json"] = json.dumps({
            b.id: {"name": b.name, "branch_id": b.university_branch_id} for b in buildings_meta
        })
        ctx["building_centers_json"] = json.dumps({
            b.id: [b.geom.centroid.y, b.geom.centroid.x]
            for b in Building.objects.exclude(geom__isnull=True)
        })
        ctx["rooms_overlay_json"] = _rooms_admin_overlay_geojson()
        ctx["edit_room_id_json"] = json.dumps(None)
        return ctx


@method_decorator(admin_required, name="dispatch")
class RoomUpdateView(AdminSaveRedirectMixin, UpdateView):
    model = Room
    form_class = RoomAdminForm
    template_name = "home/admin/room_form.html"
    list_url_name = "admin_room_list"
    add_url_name = "admin_room_add"
    edit_url_name = "admin_room_edit"

    def form_valid(self, form):
        messages.success(self.request, "Đã cập nhật phòng.")
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["buildings_json"] = buildings_geojson_dumps()
        floors_meta = Floor.objects.select_related("building", "building__university_branch").all()
        ctx["floors_meta_json"] = json.dumps({
            f.id: {
                "building_id": f.building_id,
                "building_name": (f.building.name if f.building else ""),
                "branch_id": (f.building.university_branch_id if f.building else None),
            }
            for f in floors_meta
        })
        buildings_meta = Building.objects.select_related("university_branch").all()
        ctx["buildings_meta_json"] = json.dumps({
            b.id: {"name": b.name, "branch_id": b.university_branch_id} for b in buildings_meta
        })
        ctx["building_centers_json"] = json.dumps({
            b.id: [b.geom.centroid.y, b.geom.centroid.x]
            for b in Building.objects.exclude(geom__isnull=True)
        })
        ctx["rooms_overlay_json"] = _rooms_admin_overlay_geojson()
        ctx["edit_room_id_json"] = json.dumps(self.object.pk)
        return ctx


@method_decorator(admin_required, name="dispatch")
class RoomDeleteView(DeleteView):
    model = Room
    template_name = "home/admin/confirm_delete.html"
    success_url = reverse_lazy("admin_room_list")
    context_object_name = "obj"

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["model_label"] = "Phòng"
        ctx["list_url"] = "admin_room_list"
        return ctx

    def delete(self, request, *args, **kwargs):
        messages.success(request, "Đã xóa phòng.")
        return super().delete(request, *args, **kwargs)


# ----- Building -----
@method_decorator(admin_required, name="dispatch")
class BuildingListView(AdminBulkDeleteMixin, ListView):
    model = Building
    template_name = "home/admin/building_list.html"
    context_object_name = "buildings"
    paginate_by = 10
    bulk_delete_message = "Đã xóa tòa nhà đã chọn."

    def get_queryset(self):
        qs = Building.objects.select_related("university_branch").order_by("name")
        q = (self.request.GET.get("q") or "").strip()
        if q:
            qs = qs.filter(Q(name__icontains=q) | Q(description__icontains=q) | Q(university_branch__name__icontains=q))
        return qs

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["q"] = self.request.GET.get("q", "")
        return ctx


@method_decorator(admin_required, name="dispatch")
class BuildingCreateView(AdminSaveRedirectMixin, CreateView):
    model = Building
    form_class = BuildingAdminForm
    template_name = "home/admin/building_form.html"
    list_url_name = "admin_building_list"
    add_url_name = "admin_building_add"
    edit_url_name = "admin_building_edit"

    def form_valid(self, form):
        messages.success(self.request, "Đã thêm tòa nhà mới.")
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["branch_centers_json"] = json.dumps({
            b.id: [b.geom.centroid.y, b.geom.centroid.x]
            for b in UniversityBranch.objects.exclude(geom__isnull=True)
        })
        return ctx


@method_decorator(admin_required, name="dispatch")
class BuildingUpdateView(AdminSaveRedirectMixin, UpdateView):
    model = Building
    form_class = BuildingAdminForm
    template_name = "home/admin/building_form.html"
    list_url_name = "admin_building_list"
    add_url_name = "admin_building_add"
    edit_url_name = "admin_building_edit"

    def form_valid(self, form):
        messages.success(self.request, "Đã cập nhật tòa nhà.")
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["branch_centers_json"] = json.dumps({
            b.id: [b.geom.centroid.y, b.geom.centroid.x]
            for b in UniversityBranch.objects.exclude(geom__isnull=True)
        })
        return ctx


@method_decorator(admin_required, name="dispatch")
class BuildingDeleteView(DeleteView):
    model = Building
    template_name = "home/admin/confirm_delete.html"
    success_url = reverse_lazy("admin_building_list")
    context_object_name = "obj"

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["model_label"] = "Tòa nhà"
        ctx["list_url"] = "admin_building_list"
        return ctx

    def delete(self, request, *args, **kwargs):
        messages.success(request, "Đã xóa tòa nhà.")
        return super().delete(request, *args, **kwargs)


# ----- Incident -----
@method_decorator(admin_required, name="dispatch")
class IncidentListView(AdminBulkDeleteMixin, ListView):
    model = Incident
    template_name = "home/admin/incident_list.html"
    context_object_name = "incidents"
    paginate_by = 10
    bulk_delete_message = "Đã xóa sự cố đã chọn."

    def get_queryset(self):
        qs = Incident.objects.select_related(
            "asset", "asset__equipment", "asset__tree", "incident_type",
            "building", "floor", "room"
        ).order_by("-reported_at")
        q = (self.request.GET.get("q") or "").strip()
        status = self.request.GET.get("status")
        if q:
            filters = Q(title__icontains=q) | Q(description__icontains=q)

            # Tìm theo tài sản (code thiết bị, tên thiết bị, mã cây, loài cây)
            filters |= Q(asset__equipment__code__icontains=q)
            filters |= Q(asset__equipment__name__icontains=q)
            filters |= Q(asset__tree__code__icontains=q)
            filters |= Q(asset__tree__species__icontains=q)

            # Tìm theo tên loại sự cố
            filters |= Q(incident_type__name__icontains=q)
            filters |= Q(building__name__icontains=q) | Q(floor__name__icontains=q) | Q(room__name__icontains=q)

            # Tìm theo mức độ ưu tiên (nhập chữ Thấp / Trung bình / Cao)
            q_lower = q.lower()
            matching_priorities = [
                code
                for code, label in Incident.PRIORITY_CHOICES
                if q_lower in label.lower()
            ]
            if matching_priorities:
                filters |= Q(priority__in=matching_priorities)

            qs = qs.filter(filters)
        if status:
            qs = qs.filter(status=status)
        return qs

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        for i in ctx.get("incidents", []):
            parts = []
            if i.building:
                parts.append(f"Tòa: {i.building.name}")
            if i.floor:
                parts.append(f"Tầng: {i.floor.name}")
            if i.room:
                parts.append(f"Phòng: {i.room.name}")
            i.location_summary = " | ".join(parts)
        ctx["q"] = self.request.GET.get("q", "")
        ctx["status_filter"] = self.request.GET.get("status", "")
        return ctx


def build_admin_incident_assets_detail_json() -> str:
    """Meta cho form sự cố admin: vị trí thiết bị + tọa độ cây để map zoom."""
    assets_meta_detail = {}
    for a in Asset.objects.select_related(
        "equipment",
        "equipment__room",
        "equipment__room__floor",
        "equipment__room__floor__building",
        "tree",
    ):
        if a.asset_type == "equipment" and a.equipment and a.equipment.room:
            eq_room = a.equipment.room
            eq_floor = eq_room.floor
            eq_building = eq_floor.building if eq_floor else None
            assets_meta_detail[a.id] = {
                "asset_type": "equipment",
                "equipment_id": a.equipment_id,
                "tree_id": None,
                "building_id": eq_building.id if eq_building else None,
                "floor_id": eq_floor.id if eq_floor else None,
                "room_id": eq_room.id,
                "tree_lat": None,
                "tree_lng": None,
            }
        else:
            row = {
                "asset_type": a.asset_type,
                "equipment_id": None,
                "tree_id": a.tree_id if a.asset_type == "tree" else None,
                "building_id": None,
                "floor_id": None,
                "room_id": None,
                "tree_lat": None,
                "tree_lng": None,
            }
            if a.asset_type == "tree" and a.tree and a.tree.geom:
                row["tree_lat"] = float(a.tree.geom.y)
                row["tree_lng"] = float(a.tree.geom.x)
            assets_meta_detail[a.id] = row
    return json.dumps(assets_meta_detail, ensure_ascii=False)


@method_decorator(admin_required, name="dispatch")
class IncidentCreateView(AdminSaveRedirectMixin, CreateView):
    model = Incident
    form_class = IncidentAdminForm
    template_name = "home/admin/incident_form.html"
    list_url_name = "admin_incident_list"
    add_url_name = "admin_incident_add"
    edit_url_name = "admin_incident_edit"

    def form_valid(self, form):
        messages.success(self.request, "Đã thêm sự cố mới.")
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["buildings_json"] = serialize(
            "geojson",
            Building.objects.exclude(geom__isnull=True),
            geometry_field="geom",
            fields=("name", "description"),
        )
        ctx["rooms_json"] = serialize(
            "geojson",
            Room.objects.exclude(geom__isnull=True),
            geometry_field="geom",
            fields=("name", "room_type"),
        )
        ctx["trees_json"] = serialize(
            "geojson",
            Tree.objects.exclude(geom__isnull=True),
            geometry_field="geom",
            fields=("code", "species", "health_status"),
        )
        ctx["equipment_json"] = serialize(
            "geojson",
            Equipment.objects.exclude(geom__isnull=True),
            geometry_field="geom",
            fields=("code", "name", "status"),
        )
        ctx["incidents_json"] = serialize(
            "geojson",
            Incident.objects.exclude(geom__isnull=True),
            geometry_field="geom",
            fields=("title", "status", "priority"),
        )
        ctx["incident_types_meta_json"] = json.dumps({it.id: it.applies_to for it in IncidentType.objects.all()})
        ctx["assets_meta_json"] = json.dumps({a.id: a.asset_type for a in Asset.objects.all()})
        ctx["assets_detail_json"] = build_admin_incident_assets_detail_json()
        buildings_meta = Building.objects.select_related("university_branch").all()
        floors_meta = Floor.objects.select_related("building", "building__university_branch").all()
        rooms_meta = Room.objects.select_related("floor", "floor__building", "floor__building__university_branch").all()
        ctx["buildings_meta_json"] = json.dumps({
            b.id: {"name": b.name, "branch_id": b.university_branch_id} for b in buildings_meta
        })
        ctx["floors_meta_json"] = json.dumps({
            f.id: {"name": f.name, "building_id": f.building_id, "branch_id": (f.building.university_branch_id if f.building else None)}
            for f in floors_meta
        })
        ctx["rooms_meta_json"] = json.dumps({
            r.id: {
                "name": r.name,
                "floor_id": r.floor_id,
                "building_id": (r.floor.building_id if r.floor and r.floor.building_id else None),
                "branch_id": (r.floor.building.university_branch_id if r.floor and r.floor.building else None),
            }
            for r in rooms_meta
        })
        ctx["building_centers_json"] = json.dumps({
            b.id: [b.geom.centroid.y, b.geom.centroid.x]
            for b in Building.objects.exclude(geom__isnull=True)
        })
        return ctx


@method_decorator(admin_required, name="dispatch")
class IncidentUpdateView(AdminSaveRedirectMixin, UpdateView):
    model = Incident
    form_class = IncidentAdminForm
    template_name = "home/admin/incident_form.html"
    list_url_name = "admin_incident_list"
    add_url_name = "admin_incident_add"
    edit_url_name = "admin_incident_edit"

    def form_valid(self, form):
        messages.success(self.request, "Đã cập nhật sự cố.")
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["buildings_json"] = serialize(
            "geojson",
            Building.objects.exclude(geom__isnull=True),
            geometry_field="geom",
            fields=("name", "description"),
        )
        ctx["rooms_json"] = serialize(
            "geojson",
            Room.objects.exclude(geom__isnull=True),
            geometry_field="geom",
            fields=("name", "room_type"),
        )
        ctx["trees_json"] = serialize(
            "geojson",
            Tree.objects.exclude(geom__isnull=True),
            geometry_field="geom",
            fields=("code", "species", "health_status"),
        )
        ctx["equipment_json"] = serialize(
            "geojson",
            Equipment.objects.exclude(geom__isnull=True),
            geometry_field="geom",
            fields=("code", "name", "status"),
        )
        ctx["incidents_json"] = serialize(
            "geojson",
            Incident.objects.exclude(geom__isnull=True),
            geometry_field="geom",
            fields=("title", "status", "priority"),
        )
        ctx["incident_types_meta_json"] = json.dumps({it.id: it.applies_to for it in IncidentType.objects.all()})
        ctx["assets_meta_json"] = json.dumps({a.id: a.asset_type for a in Asset.objects.all()})
        ctx["assets_detail_json"] = build_admin_incident_assets_detail_json()
        buildings_meta = Building.objects.select_related("university_branch").all()
        floors_meta = Floor.objects.select_related("building", "building__university_branch").all()
        rooms_meta = Room.objects.select_related("floor", "floor__building", "floor__building__university_branch").all()
        ctx["buildings_meta_json"] = json.dumps({
            b.id: {"name": b.name, "branch_id": b.university_branch_id} for b in buildings_meta
        })
        ctx["floors_meta_json"] = json.dumps({
            f.id: {"name": f.name, "building_id": f.building_id, "branch_id": (f.building.university_branch_id if f.building else None)}
            for f in floors_meta
        })
        ctx["rooms_meta_json"] = json.dumps({
            r.id: {
                "name": r.name,
                "floor_id": r.floor_id,
                "building_id": (r.floor.building_id if r.floor and r.floor.building_id else None),
                "branch_id": (r.floor.building.university_branch_id if r.floor and r.floor.building else None),
            }
            for r in rooms_meta
        })
        ctx["building_centers_json"] = json.dumps({
            b.id: [b.geom.centroid.y, b.geom.centroid.x]
            for b in Building.objects.exclude(geom__isnull=True)
        })
        return ctx


@method_decorator(admin_required, name="dispatch")
class IncidentDeleteView(DeleteView):
    model = Incident
    template_name = "home/admin/confirm_delete.html"
    success_url = reverse_lazy("admin_incident_list")
    context_object_name = "obj"

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["model_label"] = "Sự cố"
        ctx["list_url"] = "admin_incident_list"
        return ctx

    def delete(self, request, *args, **kwargs):
        messages.success(request, "Đã xóa sự cố.")
        return super().delete(request, *args, **kwargs)


# ----- IncidentType -----
@method_decorator(admin_required, name="dispatch")
class IncidentTypeListView(AdminBulkDeleteMixin, ListView):
    model = IncidentType
    template_name = "home/admin/incident_type_list.html"
    context_object_name = "incident_types"
    paginate_by = 10
    bulk_delete_message = "Đã xóa loại sự cố đã chọn."

    def get_queryset(self):
        qs = IncidentType.objects.all().order_by("code")
        q = (self.request.GET.get("q") or "").strip()
        if q:
            qs = qs.filter(Q(code__icontains=q) | Q(name__icontains=q) | Q(description__icontains=q))
        return qs

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["q"] = self.request.GET.get("q", "")
        return ctx


@method_decorator(admin_required, name="dispatch")
class IncidentTypeCreateView(AdminSaveRedirectMixin, CreateView):
    model = IncidentType
    form_class = IncidentTypeAdminForm
    template_name = "home/admin/incident_type_form.html"
    list_url_name = "admin_incident_type_list"
    add_url_name = "admin_incident_type_add"
    edit_url_name = "admin_incident_type_edit"

    def form_valid(self, form):
        messages.success(self.request, "Đã thêm loại sự cố mới.")
        return super().form_valid(form)


@method_decorator(admin_required, name="dispatch")
class IncidentTypeUpdateView(AdminSaveRedirectMixin, UpdateView):
    model = IncidentType
    form_class = IncidentTypeAdminForm
    template_name = "home/admin/incident_type_form.html"
    list_url_name = "admin_incident_type_list"
    add_url_name = "admin_incident_type_add"
    edit_url_name = "admin_incident_type_edit"

    def form_valid(self, form):
        messages.success(self.request, "Đã cập nhật loại sự cố.")
        return super().form_valid(form)


@method_decorator(admin_required, name="dispatch")
class IncidentTypeDeleteView(DeleteView):
    model = IncidentType
    template_name = "home/admin/confirm_delete.html"
    success_url = reverse_lazy("admin_incident_type_list")
    context_object_name = "obj"

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["model_label"] = "Loại sự cố"
        ctx["list_url"] = "admin_incident_type_list"
        return ctx

    def delete(self, request, *args, **kwargs):
        messages.success(request, "Đã xóa loại sự cố.")
        return super().delete(request, *args, **kwargs)


# ----- Asset -----
@method_decorator(admin_required, name="dispatch")
class AssetListView(AdminBulkDeleteMixin, ListView):
    model = Asset
    template_name = "home/admin/asset_list.html"
    context_object_name = "assets"
    paginate_by = 10
    bulk_delete_message = "Đã xóa tài sản đã chọn."

    def get_queryset(self):
        return Asset.objects.select_related(
            "equipment",
            "equipment__room",
            "equipment__room__floor",
            "equipment__room__floor__building",
            "equipment__room__floor__building__university_branch",
            "tree",
            "tree__university_branch",
        ).order_by("id")


@method_decorator(admin_required, name="dispatch")
class AssetCreateView(AdminSaveRedirectMixin, CreateView):
    model = Asset
    form_class = AssetAdminForm
    template_name = "home/admin/asset_form.html"
    list_url_name = "admin_asset_list"
    add_url_name = "admin_asset_add"
    edit_url_name = "admin_asset_edit"

    def form_valid(self, form):
        messages.success(self.request, "Đã thêm tài sản mới.")
        return super().form_valid(form)


@method_decorator(admin_required, name="dispatch")
class AssetUpdateView(AdminSaveRedirectMixin, UpdateView):
    model = Asset
    form_class = AssetAdminForm
    template_name = "home/admin/asset_form.html"
    list_url_name = "admin_asset_list"
    add_url_name = "admin_asset_add"
    edit_url_name = "admin_asset_edit"

    def form_valid(self, form):
        messages.success(self.request, "Đã cập nhật tài sản.")
        return super().form_valid(form)


@method_decorator(admin_required, name="dispatch")
class AssetDeleteView(DeleteView):
    model = Asset
    template_name = "home/admin/confirm_delete.html"
    success_url = reverse_lazy("admin_asset_list")
    context_object_name = "obj"

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["model_label"] = "Tài sản"
        ctx["list_url"] = "admin_asset_list"
        return ctx

    def delete(self, request, *args, **kwargs):
        messages.success(request, "Đã xóa tài sản.")
        return super().delete(request, *args, **kwargs)


# ----- Maintenance -----
@method_decorator(admin_required, name="dispatch")
class MaintenanceListView(AdminBulkDeleteMixin, ListView):
    model = Maintenance
    template_name = "home/admin/maintenance_list.html"
    context_object_name = "maintenances"
    paginate_by = 10
    bulk_delete_message = "Đã xóa phiếu bảo trì đã chọn."

    def get_queryset(self):
        qs = Maintenance.objects.select_related(
            "asset", "asset__equipment", "asset__tree", "staff"
        ).order_by("-maintenance_date")
        q = (self.request.GET.get("q") or "").strip()
        if q:
            qs = qs.filter(
                Q(asset__equipment__code__icontains=q)
                | Q(asset__equipment__name__icontains=q)
                | Q(asset__tree__code__icontains=q)
                | Q(asset__tree__species__icontains=q)
                | Q(note__icontains=q)
            )
        return qs

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["q"] = self.request.GET.get("q", "")
        return ctx


@method_decorator(admin_required, name="dispatch")
class MaintenanceCreateView(AdminSaveRedirectMixin, CreateView):
    model = Maintenance
    form_class = MaintenanceAdminForm
    template_name = "home/admin/maintenance_form.html"
    list_url_name = "admin_maintenance_list"
    add_url_name = "admin_maintenance_add"
    edit_url_name = "admin_maintenance_edit"

    def form_valid(self, form):
        messages.success(self.request, "Đã thêm phiếu bảo trì.")
        return super().form_valid(form)


@method_decorator(admin_required, name="dispatch")
class MaintenanceUpdateView(AdminSaveRedirectMixin, UpdateView):
    model = Maintenance
    form_class = MaintenanceAdminForm
    template_name = "home/admin/maintenance_form.html"
    list_url_name = "admin_maintenance_list"
    add_url_name = "admin_maintenance_add"
    edit_url_name = "admin_maintenance_edit"

    def form_valid(self, form):
        messages.success(self.request, "Đã cập nhật phiếu bảo trì.")
        return super().form_valid(form)


@method_decorator(admin_required, name="dispatch")
class MaintenanceDeleteView(DeleteView):
    model = Maintenance
    template_name = "home/admin/confirm_delete.html"
    success_url = reverse_lazy("admin_maintenance_list")
    context_object_name = "obj"

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["model_label"] = "Phiếu bảo trì"
        ctx["list_url"] = "admin_maintenance_list"
        return ctx

    def delete(self, request, *args, **kwargs):
        messages.success(request, "Đã xóa phiếu bảo trì.")
        return super().delete(request, *args, **kwargs)


# ----- Role -----
@method_decorator(admin_required, name="dispatch")
class RoleListView(ListView):
    model = Role
    template_name = "home/admin/role_list.html"
    context_object_name = "roles"
    paginate_by = 50

    def get_queryset(self):
        return Role.objects.all().order_by("name")


@method_decorator(admin_required, name="dispatch")
class RoleCreateView(AdminSaveRedirectMixin, CreateView):
    model = Role
    form_class = RoleAdminForm
    template_name = "home/admin/role_form.html"
    list_url_name = "admin_role_list"
    add_url_name = "admin_role_add"
    edit_url_name = "admin_role_edit"

    def dispatch(self, request, *args, **kwargs):
        messages.info(request, "Vai trò đang ở chế độ chỉ xem. Bạn không thể thêm mới.")
        return redirect("admin_role_list")

    def form_valid(self, form):
        messages.success(self.request, "Đã thêm vai trò mới.")
        return super().form_valid(form)


@method_decorator(admin_required, name="dispatch")
class RoleUpdateView(AdminSaveRedirectMixin, UpdateView):
    model = Role
    form_class = RoleAdminForm
    template_name = "home/admin/role_form.html"
    list_url_name = "admin_role_list"
    add_url_name = "admin_role_add"
    edit_url_name = "admin_role_edit"

    def dispatch(self, request, *args, **kwargs):
        messages.info(request, "Vai trò đang ở chế độ chỉ xem. Bạn không thể chỉnh sửa.")
        return redirect("admin_role_list")

    def form_valid(self, form):
        messages.success(self.request, "Đã cập nhật vai trò.")
        return super().form_valid(form)


@method_decorator(admin_required, name="dispatch")
class RoleDeleteView(DeleteView):
    model = Role
    template_name = "home/admin/confirm_delete.html"
    success_url = reverse_lazy("admin_role_list")
    context_object_name = "obj"

    def dispatch(self, request, *args, **kwargs):
        messages.info(request, "Vai trò đang ở chế độ chỉ xem. Bạn không thể xóa.")
        return redirect("admin_role_list")

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["model_label"] = "Vai trò"
        ctx["list_url"] = "admin_role_list"
        return ctx

    def delete(self, request, *args, **kwargs):
        messages.success(request, "Đã xóa vai trò.")
        return super().delete(request, *args, **kwargs)


# ----- AppUser -----
@method_decorator(admin_required, name="dispatch")
class AppUserListView(AdminBulkDeleteMixin, ListView):
    model = AppUser
    template_name = "home/admin/user_list.html"
    context_object_name = "users"
    paginate_by = 10
    bulk_delete_message = "Đã xóa người dùng đã chọn."

    def get_queryset(self):
        qs = AppUser.objects.select_related("role").order_by("username")
        q = (self.request.GET.get("q") or "").strip()
        if q:
            qs = qs.filter(
                Q(username__icontains=q)
                | Q(email__icontains=q)
                | Q(role__name__icontains=q)
            )
        return qs

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["q"] = self.request.GET.get("q", "")
        return ctx


@method_decorator(admin_required, name="dispatch")
class AppUserCreateView(AdminSaveRedirectMixin, CreateView):
    model = AppUser
    form_class = AppUserAdminForm
    template_name = "home/admin/user_form.html"
    list_url_name = "admin_user_list"
    add_url_name = "admin_user_add"
    edit_url_name = "admin_user_edit"

    def form_valid(self, form):
        messages.success(self.request, "Đã thêm người dùng mới.")
        return super().form_valid(form)


@method_decorator(admin_required, name="dispatch")
class AppUserUpdateView(AdminSaveRedirectMixin, UpdateView):
    model = AppUser
    form_class = AppUserAdminForm
    template_name = "home/admin/user_form.html"
    list_url_name = "admin_user_list"
    add_url_name = "admin_user_add"
    edit_url_name = "admin_user_edit"

    def form_valid(self, form):
        messages.success(self.request, "Đã cập nhật người dùng.")
        return super().form_valid(form)


@method_decorator(admin_required, name="dispatch")
class AppUserDeleteView(DeleteView):
    model = AppUser
    template_name = "home/admin/confirm_delete.html"
    success_url = reverse_lazy("admin_user_list")
    context_object_name = "obj"

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["model_label"] = "Người dùng"
        ctx["list_url"] = "admin_user_list"
        return ctx

    def delete(self, request, *args, **kwargs):
        messages.success(request, "Đã xóa người dùng.")
        return super().delete(request, *args, **kwargs)


class AdminModelTableGuardMixin:
    """
    Guard tránh crash khi DB chưa migrate bảng AboutAnnouncement.
    """

    required_table_name = ""
    required_model = None
    required_fields: list[str] = []

    def dispatch(self, request, *args, **kwargs):
        try:
            table_exists = self.required_table_name in connection.introspection.table_names()
        except (ProgrammingError, OperationalError):
            table_exists = False

        if not table_exists:
            try:
                with connection.schema_editor() as schema_editor:
                    schema_editor.create_model(self.required_model)
            except Exception:
                messages.error(
                    request,
                    "Không thể khởi tạo bảng dữ liệu tự động. "
                    "Vui lòng chạy migrate cho app home.",
                )
                return redirect("admin_dashboard")
            else:
                messages.success(
                    request,
                    "Đã tự khởi tạo bảng dữ liệu cho chức năng quản lý trang giới thiệu.",
                )
        schema_ok = self._ensure_required_columns(request)
        if not schema_ok:
            return redirect("admin_about_manage")
        return super().dispatch(request, *args, **kwargs)

    def _ensure_required_columns(self, request):
        if not self.required_fields:
            return True
        try:
            with connection.cursor() as cursor:
                description = connection.introspection.get_table_description(cursor, self.required_table_name)
            existing_cols = {col.name for col in description}
        except Exception:
            return False

        missing_fields = [name for name in self.required_fields if name not in existing_cols]
        if not missing_fields:
            return True

        for field_name in missing_fields:
            try:
                field_obj = self.required_model._meta.get_field(field_name)
                with connection.schema_editor() as schema_editor:
                    schema_editor.add_field(self.required_model, field_obj)
            except Exception:
                messages.error(
                    request,
                    f"Không thể tự bổ sung cột dữ liệu '{field_name}'. Vui lòng chạy migrate cho app home.",
                )
                return False
        return True


def _seed_about_page_sample_data() -> None:
    AboutHeroSection.objects.all().delete()
    AboutCoreValue.objects.all().delete()
    AboutAnnouncement.objects.all().delete()
    AboutFeaturedEvent.objects.all().delete()

    AboutHeroSection.objects.create(
        title="Hệ thống Quản lý Cơ sở hạ tầng Đại học",
        welcome_text=(
            "Nền tảng quản lý tập trung hỗ trợ nhà trường vận hành, giám sát và nâng cấp cơ sở hạ tầng "
            "một cách minh bạch, kịp thời và hiệu quả. Dữ liệu được cập nhật liên tục để phục vụ công tác "
            "đào tạo, nghiên cứu và trải nghiệm học tập của sinh viên."
        ),
        is_active=True,
    )

    AboutCoreValue.objects.bulk_create(
        [
            AboutCoreValue(
                title="Giảng đường thông minh",
                description="Quản lý tình trạng phòng học, thiết bị trình chiếu và năng lực sử dụng theo thời gian thực.",
                icon_class="bi bi-easel2",
                display_order=1,
                is_active=True,
            ),
            AboutCoreValue(
                title="Phòng Lab nghiên cứu",
                description="Theo dõi tài sản phòng lab, lịch bảo trì và mức độ sẵn sàng cho các hoạt động thực hành.",
                icon_class="bi bi-cpu",
                display_order=2,
                is_active=True,
            ),
            AboutCoreValue(
                title="Ký túc xá và tiện ích",
                description="Chuẩn hóa dữ liệu hạ tầng lưu trú, hệ thống điện nước và các yêu cầu bảo trì định kỳ.",
                icon_class="bi bi-building",
                display_order=3,
                is_active=True,
            ),
            AboutCoreValue(
                title="An toàn và vận hành",
                description="Tiếp nhận sự cố, cảnh báo khu vực rủi ro và hỗ trợ ra quyết định vận hành nhanh chóng.",
                icon_class="bi bi-shield-check",
                display_order=4,
                is_active=True,
            ),
        ]
    )

    AboutAnnouncement.objects.bulk_create(
        [
            AboutAnnouncement(
                title="Hoàn thành nâng cấp phòng máy tính khu A",
                summary="Nhà trường đã hoàn tất nâng cấp hệ thống máy tính và đường truyền cho khu A.",
                content=(
                    "Trong đợt bảo trì tháng này, toàn bộ phòng máy khu A đã được nâng cấp RAM, ổ cứng và "
                    "hạ tầng mạng nội bộ. Việc nâng cấp giúp tăng hiệu suất học tập, giảm thời gian chờ "
                    "khi thực hành các học phần công nghệ."
                ),
                is_published=True,
            ),
            AboutAnnouncement(
                title="Kế hoạch bảo trì thang máy tòa B",
                summary="Thang máy tòa B sẽ được bảo trì định kỳ theo kế hoạch của phòng cơ sở vật chất.",
                content=(
                    "Công tác bảo trì thang máy tòa B dự kiến diễn ra trong hai ngày cuối tuần để hạn chế "
                    "ảnh hưởng đến hoạt động giảng dạy. Trong thời gian này, sinh viên và giảng viên vui lòng "
                    "sử dụng lối thang bộ hoặc thang máy dự phòng."
                ),
                is_published=True,
            ),
            AboutAnnouncement(
                title="Bổ sung thiết bị an toàn phòng thí nghiệm",
                summary="Các phòng thí nghiệm đã được trang bị bổ sung thiết bị an toàn theo tiêu chuẩn mới.",
                content=(
                    "Nhằm nâng cao mức độ an toàn, nhà trường đã trang bị thêm bình chữa cháy, tủ dụng cụ "
                    "sơ cứu và hệ thống biển báo tại các phòng thí nghiệm trọng điểm. Danh mục thiết bị "
                    "được cập nhật đầy đủ trên hệ thống quản lý hạ tầng."
                ),
                is_published=True,
            ),
        ]
    )

    AboutFeaturedEvent.objects.bulk_create(
        [
            AboutFeaturedEvent(
                title="Lễ - Hội Đăng hương Quốc Tổ Hùng Vương & Kỷ niệm 31 năm thành lập Trường",
                summary="Sự kiện truyền thống cấp trường với nhiều hoạt động kết nối cộng đồng học thuật.",
                is_published=True,
            ),
            AboutFeaturedEvent(
                title="VLU Job Fair 2026 - Beyond Your Bound",
                summary="Ngày hội kết nối tuyển dụng với hơn 300 doanh nghiệp đồng hành.",
                is_published=True,
            ),
            AboutFeaturedEvent(
                title="Ngày hội Tìm hiểu & Trải nghiệm Mô hình Đào tạo - Future Fest Day 3",
                summary="Chuỗi trải nghiệm mô hình đào tạo mới dành cho học sinh, sinh viên và phụ huynh.",
                is_published=True,
            ),
        ]
    )


@admin_required
def admin_about_reset_sample_data(request):
    if request.method != "POST":
        return redirect("admin_about_hero_list")

    table_model_map = [
        ("home_aboutherosection", AboutHeroSection),
        ("home_aboutcorevalue", AboutCoreValue),
        ("home_aboutannouncement", AboutAnnouncement),
        ("home_aboutfeaturedevent", AboutFeaturedEvent),
    ]

    try:
        current_tables = set(connection.introspection.table_names())
    except (ProgrammingError, OperationalError):
        current_tables = set()

    for table_name, model_class in table_model_map:
        if table_name in current_tables:
            continue
        try:
            with connection.schema_editor() as schema_editor:
                schema_editor.create_model(model_class)
        except Exception:
            messages.error(
                request,
                "Không thể khởi tạo đầy đủ bảng dữ liệu trang giới thiệu. Vui lòng chạy migrate rồi thử lại.",
            )
            return redirect("admin_about_hero_list")

    _seed_about_page_sample_data()
    messages.success(request, "Đã khôi phục dữ liệu mẫu cho trang giới thiệu.")
    return redirect("admin_about_hero_list")


@method_decorator(admin_required, name="dispatch")
class AboutHeroSectionListView(AdminModelTableGuardMixin, AdminBulkDeleteMixin, ListView):
    model = AboutHeroSection
    template_name = "home/admin/about_hero_list.html"
    context_object_name = "hero_sections"
    paginate_by = 10
    bulk_delete_message = "Đã xóa Banner/Lời ngỏ đã chọn."
    required_table_name = "home_aboutherosection"
    required_model = AboutHeroSection

    def get_queryset(self):
        qs = AboutHeroSection.objects.all().order_by("-is_active", "-updated_at")
        q = (self.request.GET.get("q") or "").strip()
        if q:
            qs = qs.filter(Q(title__icontains=q) | Q(welcome_text__icontains=q))
        return qs

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["q"] = self.request.GET.get("q", "")
        return ctx


@method_decorator(admin_required, name="dispatch")
class AboutHeroSectionCreateView(AdminModelTableGuardMixin, AdminSaveRedirectMixin, CreateView):
    model = AboutHeroSection
    form_class = AboutHeroSectionAdminForm
    template_name = "home/admin/about_hero_form.html"
    list_url_name = "admin_about_hero_list"
    add_url_name = "admin_about_hero_add"
    edit_url_name = "admin_about_hero_edit"
    required_table_name = "home_aboutherosection"
    required_model = AboutHeroSection

    def form_valid(self, form):
        messages.success(self.request, "Đã tạo khối Banner & Lời ngỏ.")
        return super().form_valid(form)


@method_decorator(admin_required, name="dispatch")
class AboutHeroSectionUpdateView(AdminModelTableGuardMixin, AdminSaveRedirectMixin, UpdateView):
    model = AboutHeroSection
    form_class = AboutHeroSectionAdminForm
    template_name = "home/admin/about_hero_form.html"
    list_url_name = "admin_about_hero_list"
    add_url_name = "admin_about_hero_add"
    edit_url_name = "admin_about_hero_edit"
    required_table_name = "home_aboutherosection"
    required_model = AboutHeroSection

    def form_valid(self, form):
        messages.success(self.request, "Đã cập nhật khối Banner & Lời ngỏ.")
        return super().form_valid(form)


@method_decorator(admin_required, name="dispatch")
class AboutHeroSectionDeleteView(AdminModelTableGuardMixin, DeleteView):
    model = AboutHeroSection
    template_name = "home/admin/confirm_delete.html"
    success_url = reverse_lazy("admin_about_hero_list")
    context_object_name = "obj"
    required_table_name = "home_aboutherosection"
    required_model = AboutHeroSection

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["model_label"] = "Banner & Lời ngỏ"
        ctx["list_url"] = "admin_about_hero_list"
        return ctx

    def delete(self, request, *args, **kwargs):
        messages.success(request, "Đã xóa khối Banner & Lời ngỏ.")
        return super().delete(request, *args, **kwargs)


@method_decorator(admin_required, name="dispatch")
class AboutCoreValueListView(AdminModelTableGuardMixin, AdminBulkDeleteMixin, ListView):
    model = AboutCoreValue
    template_name = "home/admin/about_core_value_list.html"
    context_object_name = "core_values"
    paginate_by = 10
    bulk_delete_message = "Đã xóa cột giá trị cốt lõi đã chọn."
    required_table_name = "home_aboutcorevalue"
    required_model = AboutCoreValue

    def get_queryset(self):
        qs = AboutCoreValue.objects.all().order_by("display_order", "-updated_at")
        q = (self.request.GET.get("q") or "").strip()
        if q:
            qs = qs.filter(Q(title__icontains=q) | Q(description__icontains=q))
        return qs

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["q"] = self.request.GET.get("q", "")
        return ctx


@method_decorator(admin_required, name="dispatch")
class AboutCoreValueCreateView(AdminModelTableGuardMixin, AdminSaveRedirectMixin, CreateView):
    model = AboutCoreValue
    form_class = AboutCoreValueAdminForm
    template_name = "home/admin/about_core_value_form.html"
    list_url_name = "admin_about_core_value_list"
    add_url_name = "admin_about_core_value_add"
    edit_url_name = "admin_about_core_value_edit"
    required_table_name = "home_aboutcorevalue"
    required_model = AboutCoreValue

    def form_valid(self, form):
        messages.success(self.request, "Đã thêm nội dung Hiện đại & Công nghệ.")
        return super().form_valid(form)


@method_decorator(admin_required, name="dispatch")
class AboutCoreValueUpdateView(AdminModelTableGuardMixin, AdminSaveRedirectMixin, UpdateView):
    model = AboutCoreValue
    form_class = AboutCoreValueAdminForm
    template_name = "home/admin/about_core_value_form.html"
    list_url_name = "admin_about_core_value_list"
    add_url_name = "admin_about_core_value_add"
    edit_url_name = "admin_about_core_value_edit"
    required_table_name = "home_aboutcorevalue"
    required_model = AboutCoreValue

    def form_valid(self, form):
        messages.success(self.request, "Đã cập nhật nội dung Hiện đại & Công nghệ.")
        return super().form_valid(form)


@method_decorator(admin_required, name="dispatch")
class AboutCoreValueDeleteView(AdminModelTableGuardMixin, DeleteView):
    model = AboutCoreValue
    template_name = "home/admin/confirm_delete.html"
    success_url = reverse_lazy("admin_about_core_value_list")
    context_object_name = "obj"
    required_table_name = "home_aboutcorevalue"
    required_model = AboutCoreValue

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["model_label"] = "Hiện đại & Công nghệ"
        ctx["list_url"] = "admin_about_core_value_list"
        return ctx

    def delete(self, request, *args, **kwargs):
        messages.success(request, "Đã xóa nội dung Hiện đại & Công nghệ.")
        return super().delete(request, *args, **kwargs)


@method_decorator(admin_required, name="dispatch")
class AboutAnnouncementListView(AdminModelTableGuardMixin, AdminBulkDeleteMixin, ListView):
    model = AboutAnnouncement
    template_name = "home/admin/about_announcement_list.html"
    context_object_name = "announcements"
    paginate_by = 10
    bulk_delete_message = "Đã xóa thông báo đã chọn."
    required_table_name = "home_aboutannouncement"
    required_model = AboutAnnouncement
    required_fields = ["summary", "thumbnail"]

    def get_queryset(self):
        qs = AboutAnnouncement.objects.all().order_by("-is_published", "-published_at", "-created_at")
        q = (self.request.GET.get("q") or "").strip()
        status = (self.request.GET.get("status") or "").strip()
        if q:
            qs = qs.filter(Q(title__icontains=q) | Q(summary__icontains=q) | Q(content__icontains=q))
        if status == "published":
            qs = qs.filter(is_published=True)
        elif status == "hidden":
            qs = qs.filter(is_published=False)
        return qs

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["q"] = self.request.GET.get("q", "")
        ctx["status_filter"] = self.request.GET.get("status", "")
        return ctx


@method_decorator(admin_required, name="dispatch")
class AboutAnnouncementCreateView(AdminModelTableGuardMixin, AdminSaveRedirectMixin, CreateView):
    model = AboutAnnouncement
    form_class = AboutAnnouncementAdminForm
    template_name = "home/admin/about_announcement_form.html"
    list_url_name = "admin_about_announcement_list"
    add_url_name = "admin_about_announcement_add"
    edit_url_name = "admin_about_announcement_edit"
    required_table_name = "home_aboutannouncement"
    required_model = AboutAnnouncement
    required_fields = ["summary", "thumbnail"]

    def form_valid(self, form):
        messages.success(self.request, "Đã đăng thông báo mới cho trang giới thiệu.")
        return super().form_valid(form)


@method_decorator(admin_required, name="dispatch")
class AboutAnnouncementUpdateView(AdminModelTableGuardMixin, AdminSaveRedirectMixin, UpdateView):
    model = AboutAnnouncement
    form_class = AboutAnnouncementAdminForm
    template_name = "home/admin/about_announcement_form.html"
    list_url_name = "admin_about_announcement_list"
    add_url_name = "admin_about_announcement_add"
    edit_url_name = "admin_about_announcement_edit"
    required_table_name = "home_aboutannouncement"
    required_model = AboutAnnouncement
    required_fields = ["summary", "thumbnail"]

    def form_valid(self, form):
        if form.cleaned_data.get("is_published"):
            messages.success(self.request, "Đã cập nhật và đăng thông báo.")
        else:
            messages.success(self.request, "Đã cập nhật và gỡ thông báo khỏi trang giới thiệu.")
        return super().form_valid(form)


@method_decorator(admin_required, name="dispatch")
class AboutAnnouncementDeleteView(AdminModelTableGuardMixin, DeleteView):
    model = AboutAnnouncement
    template_name = "home/admin/confirm_delete.html"
    success_url = reverse_lazy("admin_about_announcement_list")
    context_object_name = "obj"
    required_table_name = "home_aboutannouncement"
    required_model = AboutAnnouncement
    required_fields = ["summary", "thumbnail"]

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["model_label"] = "Thông báo giới thiệu"
        ctx["list_url"] = "admin_about_announcement_list"
        return ctx

    def delete(self, request, *args, **kwargs):
        messages.success(request, "Đã xóa thông báo.")
        return super().delete(request, *args, **kwargs)


@method_decorator(admin_required, name="dispatch")
class AboutManageLandingView(TemplateView):
    template_name = "home/admin/about_manage_home.html"


@method_decorator(admin_required, name="dispatch")
class AboutFeaturedEventListView(AdminModelTableGuardMixin, AdminBulkDeleteMixin, ListView):
    model = AboutFeaturedEvent
    template_name = "home/admin/about_event_list.html"
    context_object_name = "events"
    paginate_by = 10
    bulk_delete_message = "Đã xóa sự kiện đã chọn."
    required_table_name = "home_aboutfeaturedevent"
    required_model = AboutFeaturedEvent

    def get_queryset(self):
        qs = AboutFeaturedEvent.objects.all().order_by("-event_date", "-published_at", "-created_at")
        q = (self.request.GET.get("q") or "").strip()
        if q:
            qs = qs.filter(Q(title__icontains=q) | Q(summary__icontains=q))
        return qs

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["q"] = self.request.GET.get("q", "")
        return ctx


@method_decorator(admin_required, name="dispatch")
class AboutFeaturedEventCreateView(AdminModelTableGuardMixin, AdminSaveRedirectMixin, CreateView):
    model = AboutFeaturedEvent
    form_class = AboutFeaturedEventAdminForm
    template_name = "home/admin/about_event_form.html"
    list_url_name = "admin_about_event_list"
    add_url_name = "admin_about_event_add"
    edit_url_name = "admin_about_event_edit"
    required_table_name = "home_aboutfeaturedevent"
    required_model = AboutFeaturedEvent

    def form_valid(self, form):
        messages.success(self.request, "Đã thêm sự kiện mới.")
        return super().form_valid(form)


@method_decorator(admin_required, name="dispatch")
class AboutFeaturedEventUpdateView(AdminModelTableGuardMixin, AdminSaveRedirectMixin, UpdateView):
    model = AboutFeaturedEvent
    form_class = AboutFeaturedEventAdminForm
    template_name = "home/admin/about_event_form.html"
    list_url_name = "admin_about_event_list"
    add_url_name = "admin_about_event_add"
    edit_url_name = "admin_about_event_edit"
    required_table_name = "home_aboutfeaturedevent"
    required_model = AboutFeaturedEvent

    def form_valid(self, form):
        messages.success(self.request, "Đã cập nhật sự kiện.")
        return super().form_valid(form)


@method_decorator(admin_required, name="dispatch")
class AboutFeaturedEventDeleteView(AdminModelTableGuardMixin, DeleteView):
    model = AboutFeaturedEvent
    template_name = "home/admin/confirm_delete.html"
    success_url = reverse_lazy("admin_about_event_list")
    context_object_name = "obj"
    required_table_name = "home_aboutfeaturedevent"
    required_model = AboutFeaturedEvent

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["model_label"] = "Sự kiện nổi bật"
        ctx["list_url"] = "admin_about_event_list"
        return ctx

    def delete(self, request, *args, **kwargs):
        messages.success(request, "Đã xóa sự kiện.")
        return super().delete(request, *args, **kwargs)
